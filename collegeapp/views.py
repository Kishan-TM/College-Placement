from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.contrib import messages
from .models import *
from openpyxl import Workbook 
import os
# Create your views here.
def index(request):
    return render(request,'index.html')



def tutor_register(request):
    if request.method=='POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        pswd = request.POST.get('pswd')  # Password input (ensure it is hashed before saving)
        date_of_birth = request.POST.get('date_of_birth')
        phone = request.POST.get('phn')
        address = request.POST.get('address')
        experience_years = request.POST.get('experience_years')
        subject = request.POST.get('subject')
        qualification = request.POST.get('qualification')
        resume = request.FILES.get('resume')
        Tutor(
            name=name,
            email=email,
            pswd=pswd,
            date_of_birth=date_of_birth,
            phone=phone,
            address=address,
            experience_years=experience_years,
            subject=subject,
            qualification=qualification,
            resume=resume,
        ).save()
        txt="""<script>alert("Registration successfull ");window.location='/login/';</script>"""
        return HttpResponse(txt)
    return render(request,'tutor/tutor_register.html')

def add_student(request):
    if request.method=='POST':
        name=request.POST.get('name')
        email=request.POST.get('email')
        pswd=request.POST.get('pswd')
        course=request.POST.get('course')
        # student_class=request.POST.get('student_class')
        Student(
            name=name,
            email=email,
            pswd=pswd,
            course=course,
            # student_class=student_class
        ).save()
        txt="""<script>alert("Student Added successfully ");window.location='/add_student/';</script>"""
        return HttpResponse(txt)
    return render(request,'tutor/add_student.html')


def login(request):
    if request.method=='POST':
        email=request.POST.get('email')
        pswd=request.POST.get('pswd')

        student=Student.objects.filter(email=email, pswd=pswd).first()
        tutor=Tutor.objects.filter(email=email, pswd=pswd).first()

        if student:
            request.session['id'] = student.id
            request.session['role'] = 'student'
            txt = """<script>alert('Login successful!'); window.location='/studentindex/';</script>"""
            return HttpResponse(txt)
        elif tutor:
            # Only allow login if the tutor is approved
            if tutor.is_approved:
                request.session['id'] = tutor.id
                request.session['role'] = 'tutor'
                txt = """<script>alert('Login successful!'); window.location='/tutorindex/';</script>"""
                return HttpResponse(txt)
            else:
                # Tutor not approved
                txt = """<script>alert('Your account is not approved yet. Please contact the admin.'); window.location='/login/';</script>"""
                return HttpResponse(txt)
            
        elif email == 'admin@gmail.com' and pswd == 'admin':
             request.session['role'] = 'admin'
             txt = """<script>alert('Login successful!'); window.location='/adminindex/';</script>"""
             return HttpResponse(txt)

        txt = """<script>alert('Invalid credentials, please try again!'); window.location='/login/';</script>"""
        return HttpResponse(txt)

    return render(request, 'login.html')

def studentindex(request):
    return render(request,'studentindex.html')

def tutorindex(request):
    return render(request,'tutorindex.html')

def adminindex(request):
    return render(request,'adminindex.html')

def student_list(request):
    tid=request.session.get('id')
    tutor=Tutor.objects.get(id=tid)
    students=Student.objects.filter(course=tutor.subject)
    return render(request,'tutor/student_list.html',{'students':students})

def view_profile(request):
    sid=request.GET.get('id')
    student=Student.objects.filter(id=sid).first()
    return render(request, 'tutor/view_profile.html', {'student': student})

def view_student_list(request):
    students=Student.objects.all()
    return render(request,'admin/view_student_list.html',{'students':students})

def view_student_profile(request):
    sid=request.GET.get('id')
    student=Student.objects.filter(id=sid).first()
    return render(request, 'admin/view_student_profile.html', {'student': student})

def tutor_lists(request):
    return render(request,'admin/tutor_lists.html')

def manage_tutor(request):
    tutors=Tutor.objects.all()
    return render(request,'admin/manage_tutor.html',{'tutors':tutors})

def approve_tutor(request):
    tutor_id=request.GET.get('id')
    tutor=Tutor.objects.get(id=tutor_id)
    tutor.is_approved = True
    tutor.status = "Approved"
    tutor.save()
    messages.success(request, f"Tutor {tutor.name} has been successfully approved!")  # Add success message
    return redirect('tutor_lists') 

def reject_tutor(request):
    tutor_id=request.GET.get('id')
    tutor=Tutor.objects.get(id=tutor_id)
    tutor.is_approved= False
    tutor.status = "Rejected"
    tutor.save()
    messages.success(request, f"Tutor {tutor.name} has been successfully rejected.")
    return redirect('tutor_lists')

def create_job_list(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        qualifications = request.POST.get('qualifications')
        location = request.POST.get('location')
        company_name = request.POST.get('company_name')
        is_active = request.POST.get('is_active') == 'on'

        JobListing.objects.create(
            title=title,
            description=description,
            qualifications=qualifications,
            location=location,
            company_name=company_name,
            is_active=is_active
        )
        txt = """<script>alert('Job Listing Created Successfully!'); window.location='/create_job_list/';</script>"""
        return HttpResponse(txt)
    return render(request, 'admin/create_job_list.html')

def manage_job_list(request):
    job_list=JobListing.objects.all()
    return render(request,'admin/manage_job_list.html',{'job_listings': job_list})

def edit_job_list(request):
    job_id=request.GET.get('id')
    job=JobListing.objects.filter(id=job_id).first( )
    if request.method == 'POST':
        job.title = request.POST.get('title')
        job.description = request.POST.get('description')
        job.qualifications = request.POST.get('qualifications')
        job.location = request.POST.get('location')
        job.company_name = request.POST.get('company_name')
        job.is_active = request.POST.get('is_active') == 'on'  # Checkbox value to boolean
        job.save()
        return redirect('manage_job_list')

    return render(request, 'admin/edit_job_list.html', {'job': job})

def delete_training_session(request):
    session_id=request.GET.get('id')
    sessions=TrainingSession.objects.filter(id= session_id).first()
    sessions.delete()
    return redirect('training_session_list')

def delete_job_list(request):
    job_id=request.GET.get('id')
    job=JobListing.objects.filter(id=job_id).first()
    job.delete()
    return redirect('manage_job_list')

def approved_tutors(request):
    tutors = Tutor.objects.filter(is_approved=True)
    return render(request, 'admin/approved_tutors.html', {'tutors': tutors})

def rejected_tutors(request):
    tutors=Tutor.objects.filter(is_approved=False)
    return render(request,'admin/rejected_tutors.html',{'tutors':tutors})

def available_jobs(request):
    active_jobs = JobListing.objects.filter(is_active=True).order_by('-created_at')
    return render(request, 'tutor/available_jobs.html', {'jobs': active_jobs})
    
def student_job_list(request):
    student_id = request.session.get('id')
    student=Student.objects.get(id=student_id)
    job=JobListing.objects.filter(course=student.course)
    return render(request,'student/student_job_list.html',{'jobs':job})

    
def student_list(request):
    tid=request.session.get('id')
    tutor=Tutor.objects.get(id=tid)
    students=Student.objects.filter(course=tutor.subject)
    return render(request,'tutor/student_list.html',{'students':students})


def update_profile(request):
    student_id = request.session.get('id')
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        raise Http404("Student not found")
    if request.method == 'POST':
        student.name = request.POST.get('name', student.name)
        student.email = request.POST.get('email', student.email)
        student.pswd = request.POST.get('pswd', student.pswd)
        student.date_of_birth = request.POST.get('date_of_birth', student.date_of_birth)
        student.phn = request.POST.get('phn', student.phn)
        student.address = request.POST.get('address', student.address)
        student.student_class = request.POST.get('student_class', student.student_class)
        student.course = request.POST.get('course', student.course)
        student.supply = request.POST.get('supply', student.supply)
        
        
        if 'resume' in request.FILES:
            student.resume = request.FILES['resume']
        
        if 'picture' in request.FILES:
            student.picture = request.FILES['picture']
        
        student.save()

        return redirect('update_profile')
    
    return render(request, 'student/update_profile.html', {'student': student})

def student_profile(request):
    student_id = request.session.get('id')
    print("Student ID from GET:", student_id)
    student = Student.objects.filter(id=student_id).first()
    return render(request, 'student/student_profile.html', {'student': student})


def verify_supplies(request):
    # Get the tutor's ID from the session
    tid = request.session.get('id')

    # Check if tutor exists
    try:
        tutor = Tutor.objects.get(id=tid)
    except Tutor.DoesNotExist:
        raise Http404("Tutor not found")

    # Filter students who have supplies (supply is not null) and whose course matches the tutor's course
    students = Student.objects.filter( course=tutor.subject)

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        verify_status = request.POST.get('verified') == 'True'

        # Retrieve the student whose supply is not null and whose course matches the tutor's course
        try:
            student = Student.objects.get(id=student_id, course=tutor.subject )
        except Student.DoesNotExist:
            return redirect('verify_supplies')

        student.verified = verify_status
        student.save()
        return redirect('verify_supplies')

    return render(request, 'tutor/verify_supplies.html', {'students': students})

def apply_for_job(request):
    student_id = request.session.get('id')  # Get the student ID from session
    job_id = request.GET.get('id')  # Get the job ID from the URL

    # Check if the student is logged in
    if not student_id:
        messages.error(request, "You must be logged in to apply for a job.")
        return redirect('login')

    try:
        student = Student.objects.get(id=student_id)  # Fetch the student object
    except Student.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('login')  # Redirect to the login page if the student doesn't exist
    # Check if the student is verified
    if not student.verified:
        txt = """<script>alert(' "You must be a verified student to apply for a job.'); window.location='/studentindex/';</script>"""
        return HttpResponse(txt)
    job = JobListing.objects.filter(id=job_id).first()
    # Check if the job exists
    if not job:
        messages.error(request, "Job not found.")
        return redirect('student_job_list')
    if request.method == 'POST':
        JobApplication.objects.create(student=student, job=job)
        txt = """<script>alert('"You have successfully applied for the job!'); window.location='/student_job_list/';</script>"""
        return HttpResponse(txt)
    return render(request, 'student/apply_for_job.html', {'job': job, 'student': student})

def view_student_applications(request):
    tutor_id=request.session.get('id')
    tutor=Tutor.objects.get(id=tutor_id)
    applications = JobApplication.objects.filter(student__course=tutor.subject)
    return render(request, 'tutor/view_student_applications.html', {'applications': applications})

def track_applications(request):
    student_id=request.session.get('id')
    student=Student.objects.get(id=student_id)
    applications = JobApplication.objects.filter(student=student)
    return render(request, 'student/track_applications.html', {'applications': applications})

def tutor_profile(request):
    tutor_id=request.session.get('id')
    tutor=Tutor.objects.filter(id=tutor_id).first()
    return render(request,'tutor/tutor_profile.html',{'tutor':tutor})

def tutor_edit_profile(request):
    tutor_id = request.session.get('id')
    try:
        tutor = Tutor.objects.get(id=tutor_id)
    except Tutor.DoesNotExist:
        raise Http404("Student not found")
    if request.method == 'POST':
        tutor.name = request.POST.get('name', tutor.name)
        tutor.email = request.POST.get('email', tutor.email)
        tutor.pswd = request.POST.get('pswd', tutor.pswd)
        tutor.date_of_birth = request.POST.get('date_of_birth', tutor.date_of_birth)
        tutor.experience_years = request.POST.get('experience_years', tutor.experience_years)
        tutor.address = request.POST.get('address', tutor.address)
        tutor.qualification = request.POST.get('qualification', tutor.qualification)
        tutor.subject = request.POST.get('subject', tutor.subject)
    
        if 'resume' in request.FILES:
            tutor.resume = request.FILES['resume']
        
        tutor.save()
        return redirect('tutor_edit_profile')
    
    return render(request, 'tutor/tutor_edit_profile.html', {'tutor': tutor})

def add_training_session(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        date = request.POST.get('date')
        location = request.POST.get('location')
        instructor = request.POST.get('instructor')

        session = TrainingSession(
            title=title,
            description=description,
            date=date,
            location=location,
            instructor=instructor,
        )
        session.save()
        txt = """<script>alert('You have successfully added a new training session!'); window.location='/training_session_list/';</script>"""
        return HttpResponse(txt)
    return render(request, 'admin/add_training_session.html')

def training_session_list(request):
    sessions = TrainingSession.objects.all()
    return render(request, 'admin/training_session_list.html', {'sessions': sessions})

def student_training_session(request):
    sessions= TrainingSession.objects.all()
    return render(request,'student/student_training_session.html',{'sessions':sessions})

def enroll_training_session(request):
    student_id = request.session.get('id')
    session_id = request.GET.get('id')

    student = Student.objects.filter(id=student_id).first()
    session = TrainingSession.objects.filter(id=session_id).first()

    if not student or not session:
        messages.error(request, "Invalid student or training session.")
        return redirect('student_training_session')

    if request.method == 'POST':
        if session.attendees.filter(id=student.id).exists():
            txt = """<script>alert('You are already enrolled.'); window.location='/student_training_session/';</script>"""
            return HttpResponse(txt)
        else:
            session.attendees.add(student)
            txt = """<script>alert('Enrollment successful.'); window.location='/student_training_session/';</script>"""
            return HttpResponse(txt)
    return render(request, 'student/enroll_training_session.html', {'session': session})



def export_training_session_attendees(request):
    session_id= request.GET.get('id')
    session=TrainingSession.objects.get(id=session_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendees for {session.title}"

    ws.append(['Name', 'Email', 'Phone', 'Course', 'Class', 'Resume', 'Verified'])

    for student in session.attendees.all():
        resume_file_name = os.path.basename(student.resume.name) if student.resume else 'No Resume'
        ws.append([
            student.name, 
            student.email, 
            student.phn, 
            student.course, 
            student.student_class, 
            resume_file_name,  # Display the file name
            student.verified
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendees_{session.title}.xlsx"'

    wb.save(response)
    return response

def application_list(request):
    job_applications=JobApplication.objects.all()
    return render(request,'admin/application_list.html',{'job_applications': job_applications})

# def export_job_applications(request):
#     job_applications = JobApplication.objects.select_related('student', 'job').all()

#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Job Applications'

#     ws.append(['Student Name', 'Student Email', 'Job Title', 'Application Date', 'Status'])

#     for application in job_applications:
#         ws.append([
#             application.student.name, 
#             application.student.email, 
#             application.job.title, 
#             application.application_date,
#             application.status
#         ])
#     response = HttpResponse(
#     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="job_applications.xlsx"'

#     wb.save(response)
#     return response

from django.http import HttpResponse
from openpyxl import Workbook

def export_job_applications(request):
    job_applications = JobApplication.objects.select_related('student', 'job').all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Job Applications'

    # Add headers
    ws.append(['Student Name', 'Student Email', 'Job Title', 'Application Date', 'Status'])

    # Add data rows
    for application in job_applications:
        ws.append([
            application.student.name,
            application.student.email,
            application.job.title,
            application.application_date.strftime('%Y-%m-%d %H:%M'),
            application.status
        ])

    # Set up HTTP response for Excel file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="job_applications.xlsx"'

    wb.save(response)
    return response

def export_students_to_excel(request):
    tutor_id=request.session.get('id')
    if not tutor_id:
        return HttpResponse("Tutor not logged in.", status=403)
    
    try:
        tutor = Tutor.objects.get(id=tutor_id)
    except Tutor.DoesNotExist:
        return HttpResponse("Tutor not found.", status=404)

    

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Details"

    ws.append(["Name", "Email", "Date of Birth", "Phone", "Address", "Class", "Course", "Supply", "Verified"])

    students = Student.objects.filter(course=tutor.subject)
    for student in students:
        ws.append([
            student.name,
            student.email,
            student.date_of_birth.strftime('%Y-%m-%d %H:%M') if student.date_of_birth else "N/A",
            student.phn if student.phn else "N/A",
            student.address if student.address else "N/A", 
            student.student_class if student.student_class else "N/A",
            student.course if student.course else "N/A",
            student.supply if student.supply else "N/A",
            "Yes" if student.verified else "No",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'

    # Save workbook to the response
    wb.save(response)
    return response